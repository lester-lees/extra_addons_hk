# -*- encoding: utf-8 -*-
from openerp.osv import fields,osv
from openpyxl.reader.excel import load_workbook 
from openerp import tools
from openerp.tools.translate import _
import os
import re
import datetime
import logging
_logger = logging.getLogger(__name__)
	
class product_tmalljd_gifts(osv.osv):
    _name = "product.tmalljd.gifts"
	
    _columns = {
        'tmalljd_id': fields.many2one('product.tmalljd', 'TMallJD ID'),
        'product_id': fields.many2one('product.product','ERP Product Name'),
        'price': fields.float('Price'),		
        'qty': fields.integer('Quantity'),	
    }

class product_tmalljd(osv.osv):
    _name = "product.tmalljd"
    #_inherits = {'product.product': 'product_id'}

    def _get_ean13(self, cr, uid, ids, field_name, arg, context=None):	
        result = {}
        for line in self.pool.get('product.tmalljd').browse(cr, uid, ids, context=context):
            if line.erp_product_id : result[line.id] = line.erp_product_id.ean13 or line.erp_product_id.default_code
        return result		

    def _get_stock(self, cr, uid, ids, field_name, arg, context=None):	
        result = {}
        domain_products = [('location_id','=',38)]		
        quants = self.pool.get('stock.quant').read_group(cr, uid, domain_products, ['product_id', 'qty'], ['product_id'], context=context)	
        quants = dict(map(lambda x: (x['product_id'][0], x['qty']), quants))
		
        for line in self.pool.get('product.tmalljd').browse(cr, uid, ids, context=context):
            id = line.id
            if line.erp_product_id :
                pid = line.erp_product_id.id			
                result[id] = quants.get(pid, 0.0)
            else:
                result[id] = 0			
        return result		
		
    _columns = {
        'gift_ids': fields.one2many('product.tmalljd.gifts','tmalljd_id','Gift Products'), 
        'erp_product_set': fields.many2many('product.product', 'product_product_tmalljd_rel', 'tmalljd_id', 'product_id', u'Product Set(套装)'),	
        'erp_product_id': fields.many2one('product.product','ERP Name'),
        'erp_ean13': fields.char('ERP_EAN13'), #fields.function(_get_ean13,type='char',string='ERP_EAN13'),
        'erp_stock': fields.float('ERP_Stock'),#fields.function(_get_stock,type='float',string='ERP库存'),		
        'ec_shop_id': fields.many2one('loewieec.shop', u'店铺'),		
        'ec_num_iid': fields.char(u'电商数字编码'),
        'ec_sku_id': fields.char(u'SKU编码'),
        'ec_title':fields.char(u'商品标题'),
        'ec_price':fields.float(u'售价'),
        'ec_color':fields.char(u'颜色'),
        'ec_ean13': fields.char(u'条形码'),
        'ec_brand': fields.char(u'品牌'),
        'ec_qty': fields.integer(u'EC数量'),
        'ec_outer_code': fields.char(u'商家外部编码'),		
        'ec_product_name': fields.char(u'产品名称'),	
        'ec_product_id': fields.char(u'EC产品ID'),		
        'ec_num_custom':fields.char(u'海关代码'),	
    }

	
class loewieec_error(osv.osv):
    _name = "loewieec.error"

    _columns = {	
        'shop_id': fields.many2one('loewieec.shop', u'店铺'),
        'name': fields.char(u'错误信息'),
    }

	
class sale_order_line(osv.osv):
    _inherit = "sale.order.line"
	
    _columns = {
        'logistic_sent': fields.related('coe_no', 'logistic_sent', type='boolean', string=u'已同步运单?',readonly=True),	
        'coe_no': fields.many2one('sale.coe',string=u'COE单号'),	
        'tmi_jdi_no': fields.char(string=u'电商单号'),
        'buyer_nick': fields.char(u'买家昵称'),		
        'pay_time': fields.datetime(u'EC支付时间'),	
        'create_time_tmjd': fields.datetime(u'EC创建时间'),		
    }	
	
    def copy_sale_order_line(self, cr, uid, ids, context=None): 
        for line in self.pool.get('sale.order.line').browse(cr, uid, ids, context=context):
            line.copy()	

	
class sale_order(osv.osv):
    _name = "sale.order"
    _inherit = "sale.order"
        
    _columns = {	
        'express_ids': fields.related('order_line', 'coe_no', type='many2one', relation='sale.coe', string=u'TMI_JDI收货人'),	
        'tmi_jdi_nos': fields.related('order_line', 'tmi_jdi_no', type='char', string='TMI_JDI_NO'),	
        'selected': fields.boolean('Selected'),	
        'shop_id': fields.many2one('loewieec.shop', string=u"EC店铺名", readonly=True),
        'sale_code': fields.char(u'EC单号', readonly=True),	
        'tid': fields.char(u'交易单号', readonly=True),	
        'buyer_nick': fields.char(u'买家昵称'),			
        'order_state': fields.selection([
            ('WAIT_SELLER_SEND_GOODS', u'等待卖家发货'),
            ('WAIT_BUYER_CONFIRM_GOODS', u'等待买家确认收货'),
            ('TRADE_FINISHED', u'交易成功'),
            ('TRADE_CLOSED', u'交易关闭'),
            ], u'订单状态'),
    }

    def update_orders_seller_memo(self, cr, uid, ids, context=None):
        sale_order_obj = self.pool.get('sale.order').browse(cr,uid,ids[0],context=context)	
        shop = sale_order_obj.shop_id		
        if not shop : return False 
        if shop.code == 'JDI' :
            raise osv.except_osv(u'错误',u'''JDI京东国际订单无需更新备注''') 
            return False			
			
        statement = "select tmi_jdi_no from sale_order_line where order_id=%d group by tmi_jdi_no" % ids[0] 
        cr.execute(statement)
        tids = [item[0] for item in cr.fetchall()]
        if not tids : return False
		
        return shop.update_orders_seller_memo(context=context, tids=tids)		
	
    def delete_lines_of_tmijdi_no(self, cr, uid, ids, context=None):   # 完整删除 天猫京东 订单的 行
        sale_order_obj = self.pool.get('sale.order').browse(cr,uid,ids[0],context=context)

        note = 	sale_order_obj.note or ''
        tmijdi_nos = note.strip().split(',') 		
        tmijdi_no_list = []
	
        for tmijdi_no in tmijdi_nos:
            if tmijdi_no.strip() != '': tmijdi_no_list.append( tmijdi_no.strip() )
			
        statement = "delete from sale_order_line where order_id=%d and tmi_jdi_no in (%s)" %  ( ids[0], ("'" + """','""".join(tmijdi_no_list)	+ "'") )		
        cr.execute(statement)				

        val = val1 = 0.0
        cur = sale_order_obj.pricelist_id.currency_id
        for line in sale_order_obj.order_line:
            val1 += line.price_subtotal
            val += self._amount_line_tax(cr, uid, line, context=context)
			
        cur_obj = self.pool.get('res.currency')			
        amount_tax = cur_obj.round(cr, uid, cur, val)		
        amount_untaxed = cur_obj.round(cr, uid, cur, val1)		
        amount_total = amount_untaxed + amount_tax		
        sale_order_obj.write({'amount_tax':amount_tax, 'amount_untaxed': amount_untaxed,'amount_total':amount_total})
				
    def delete_multi_gift_lines(self, cr, uid, ids, context=None):
        sale_order_obj = self.pool.get('sale.order').browse(cr,uid,ids[0],context=context)
        gift_product_id = sale_order_obj.shop_id.gift_product_id.id			
        coe_list = []
        delete_list = []
		
        for line in sale_order_obj.order_line.filtered(lambda r: r.product_id.id == gift_product_id):		
            if line.coe_no.name in coe_list : 
                delete_list.append( line.coe_no.name )			
                line.unlink()
            else:
                coe_list.append(line.coe_no.name)			

        if delete_list :
            log = sale_order_obj.note or ''		
            sale_order_obj.note = u"删除了以下运单号的重复赠品行：" + chr(10) + ','.join(delete_list) + chr(10) + log 		
		
    def delete_no_coeno_lines(self, cr, uid, ids, context=None):
	
        statement = "(select s.id from sale_order_line s left join sale_coe c on s.coe_no=c.id where s.order_id=%d and trim(c.name) not like 'EL%sHK') union (select id from sale_order_line where order_id=%d and coe_no is Null)" % (ids[0], '%', ids[0])	
        cr.execute(statement)		
        line_ids = [ item[0] for item in cr.fetchall() ]		
        sale_line_obj = self.pool.get('sale.order.line').unlink(cr,uid, line_ids,context=context)       		
		
    def update_waybill_no(self, cr, uid, ids, context=None):
        sale_order_obj = self.pool.get('sale.order').browse(cr,uid,ids[0],context=context)	
        shop = sale_order_obj.shop_id		
        if not shop : return False 
        if shop.code == 'JDI' :
            return shop.jdi_order_delivery(salesorder=sale_order_obj, context=context)  		
        
        return shop.update_tmall_waybill(context=context, salesorder=sale_order_obj)			
	
    def view_express_data(self, cr, uid, ids, context=None):	
        sale_order_obj = self.pool.get('sale.order').browse(cr,uid,ids[0],context=context)		
        if not sale_order_obj : 
            raise osv.except_osv(u'Sale order 错误',u'''请先保存销售单草稿''')	
            return False

        sale_order_line_ids = self.pool.get('sale.order.line').search(cr,uid,[('order_id','=',ids[0])],context=context)	
        if len(sale_order_line_ids)< 1: return False		
        eids = self.pool.get('sale.order.line').read(cr,uid,sale_order_line_ids,['coe_no'],context=context)	
        express_ids = [ eid['coe_no'] and eid['coe_no'][0] for eid in eids ]			
		
        customer_id = sale_order_obj.partner_id.id
        sale_coe_obj = self.pool.get('sale.coe')
        platform = sale_order_obj.shop_id.code		
        if len(express_ids)>0:		
            for express_obj in sale_coe_obj.browse(cr,uid,express_ids,context=context):
                express_obj.sale_id = ids[0]		
                express_obj.customer = customer_id	                		
		
        mod_obj = self.pool.get('ir.model.data')
        act_obj = self.pool.get('ir.actions.act_window')

        result = mod_obj.get_object_reference(cr, uid, 'loewieec_sync_hk', 'action_loewieec_salecoe')
        id = result and result[1] or False
        result = act_obj.read(cr, uid, [id], context=context)[0]
        result['domain'] = [('id','in',express_ids)]
        result['res_id'] = express_ids
        result['context'] = {'default_sale_id':ids[0],'default_customer':customer_id}		
		
        return result

    def split_address(self, address):
        assert address != None
        addr_list = address.split(" ",2)		
        if len(addr_list) < 3: return None	
		
        length = len( addr_list[2] )
        for i in range( length-1, 0, -1 ):
            if addr_list[2][i] == '(': break

        addr = addr_list[2][0:i]
        zip = addr_list[2][i+1: length-1]		
		
        return {'province':addr_list[0],'city':addr_list[1],'address':addr,'zip':zip}				
		
    def get_full_path(self, cr, uid, path):
        # sanitize ath
        path = re.sub('[.]', '', path)
        path = path.strip('/\\')
        return os.path.join(tools.config.filestore(cr.dbname), path)	
		
    def import_tmall_csv_file(self, cr, uid, ids, context=None):

        attachment_obj = self.pool.get('ir.attachment')
        shop_attach_id = context.get('res_id')	
        shop_attach_id = shop_attach_id or ids[0]		
        attachment_id = attachment_obj.search(cr,uid,[('res_id', '=', shop_attach_id)], context=context)		
        if len(attachment_id)<1: return False
        attach = attachment_obj.browse(cr,uid,attachment_id[0],context=context)
        fname = attach.store_fname
        display_name = attach.name		
        if not fname : return False		
        fname = self.get_full_path(cr, uid, fname)	
		
        csvfile = file(fname, 'rb')		
		
        coe_obj = self.pool.get('sale.coe')    		
        tmi_no = {}
        no_coe_list = []		
        		
        statement = 'select tmi_jdi_no from sale_order_line where order_id=%d group by tmi_jdi_no' % ids[0]
        cr.execute(statement)
        tmi_jdi_no_list = [ str(item[0]).strip() for item in cr.fetchall() ]		
				
        sale_id = ids[0]
        #for line in reader:
        for cols in csvfile.readlines():		
            cols = cols.decode('gbk')
            cols = cols.split(",")	
            line = [ col.replace('"','').strip() for col in cols ]
            line[0] = line[0].replace('=','').strip()			
            if line[0] == u'订单编号' and line[1] == u'买家会员名' or line[0] not in tmi_jdi_no_list :	 continue

            #coeids = coe_obj.search(cr,uid,[('name','=',line[12]),('address','=',addr and addr2 or line[13]),('tel','=',phone)],context=context)
            coeids = coe_obj.search(cr,uid,[('tmi_jdi_no','=',line[0])],context=context)			
            coeid = coeids and coeids[0] or 0				
            if not coeid : 	
			
                addr = self.split_address(line[13])		
                addr2 = addr['province'] + ',' + addr['city'] + ',' + addr['address']
                phone = line[16].replace("'","")
                phone = phone.strip()
			
                if addr : coe_info = {'tmi_jdi_no':line[0], 'name':line[12], 'receive_name':line[12], 'mobile':phone, 'tel':phone, 'address':addr2, 'province':addr['province'], 'city':addr['city'], 'zip':addr['zip']}
                else: coe_info = {'tmi_jdi_no':line[0], 'name':line[12], 'receive_name':line[12], 'mobile':phone, 'tel':phone, 'address':line[13]}				
                coeid = coe_obj.create( cr, uid, coe_info, context=context )
 
            tmi_no[line[0]] = coeid
			
        sale_order = self.pool.get('sale.order').browse(cr,uid,ids[0],context=context)
        coe_in_sale = []		
        for line in sale_order.order_line:
            tmi_jdi_no = line.tmi_jdi_no and line.tmi_jdi_no.strip() or ''
            if not tmi_jdi_no : continue
			
            if tmi_jdi_no in tmi_no.keys():
                line.write( {'express_id': tmi_no[tmi_jdi_no]} )  # 深圳ERP中为 line.express_id 而香港ERP为 line.coe_no
                coe_in_sale.append(tmi_jdi_no)	
                #_logger.info( "Jimmy: have such a order:%s, and coeid:%d " % ( tmi_jdi_no, tmi_no[tmi_jdi_no]) )				
                #tmi_no[tmi_jdi_no] = 0	 # 这里有问题， 如果一个 电商订单有几个产品，则只有第一个产品行 会被添加 coe 条目id			
            else:			
                no_coe_list.append(tmi_jdi_no) 

        if no_coe_list : 
            log = sale_order.note or ''		
            sale_order.note = u'以下销售订单行的TMI_NO在CSV文件中不存在:' + chr(10) + ",".join(no_coe_list) + chr(10) + log
			
        not_in_tmi_no = []
        for key in tmi_no.keys():
            if key not in  coe_in_sale: not_in_tmi_no.append(key)
			
        if not_in_tmi_no : 
            log = sale_order.note or ''		
            sale_order.note = u"以下CSV内的 '订单编号' 无法匹配到 销售订单行:" + chr(10) + ",".join( not_in_tmi_no ) # + chr(10) + log	  #
			
        #_logger.info( "Jimmy: OK....")			
        return True		
		
	
class sale_coe(osv.osv):
    _name = "sale.coe"	
	
    _columns = {	
        'logistic_sent': fields.boolean(u'已同步运单?',default=False, readonly=True, copy=False),	
        'sale_id': fields.many2one('sale.order', string='Sales Order', readonly=True, states={'draft': [('readonly', False)]} , copy=False),
        'picking_id': fields.many2one('stock.picking',string='Picing Order', readonly=True, states={'draft': [('readonly', False)]}, copy=False),		
        'customer': fields.many2one('res.partner',string=u'客户', readonly=True, states={'draft': [('readonly', False)]}, copy=False),		
        'tmi_jdi_no': fields.char(string='TMI JDI NO', readonly=True, states={'draft': [('readonly', False)]}),	
		
        'name':fields.char(string='COE NO', readonly=True, states={'draft': [('readonly', False)]}),	
        'receive_name': fields.char(string='Receive Name', readonly=True, states={'draft': [('readonly', False)]}),		
        'tel': fields.char(string='Cell Phone', readonly=True, states={'draft': [('readonly', False)]}),
        'telephone': fields.char(string='Telephone', readonly=True, states={'draft': [('readonly', False)]}),		
        'province': fields.char(string='Province', readonly=True, states={'draft': [('readonly', False)]}),	
        'city': fields.char(string='City', readonly=True, states={'draft': [('readonly', False)]}),	
        'county': fields.char(string='County', readonly=True, states={'draft': [('readonly', False)]}),			
        'address': fields.char(string='Address', readonly=True, states={'draft': [('readonly', False)]}),	
        'zip': fields.char(string='Zip', readonly=True, states={'draft': [('readonly', False)]}),
		
        'class_desc': fields.char(string='Desc',default=u'None', readonly=True, states={'draft': [('readonly', False)]}),
        'qty': fields.integer(string='Quantity', default=1, readonly=True, states={'draft': [('readonly', False)]}),		
        'price': fields.float(string='Fee',default=50, readonly=True, states={'draft': [('readonly', False)]}),	
        'weight': fields.float(string='Weight',default=0.2, readonly=True, states={'draft': [('readonly', False)]}),
		
        'state': fields.selection([('draft',u'草稿'),('done',u'完成')],string='State',default='draft'),		
    }

			
class stock_move(osv.osv):
    _inherit = "stock.move"
	
    def _get_coe_no(self, cr, uid, ids, field_name, arg, context=None):	
        result = {}
        for move in self.pool.get('stock.move').browse(cr, uid, ids, context=context):
            result[move.id] = move.procurement_id.sale_line_id.coe_no.id
        return result	
		
    _columns = {
        #'sale_order_line': fields.function(_get_sale_order_line, type='char',string='Sales Line'),	
        'coe_no': fields.function(_get_coe_no,type='many2one',relation='sale.coe',string='COE NO'),
    }		
	
	
class stock_picking(osv.osv):
    _inherit = "stock.picking"
	
    def get_full_path(self, cr, uid, path):
        # sanitize ath
        path = re.sub('[.]', '', path)
        path = path.strip('/\\')
        return os.path.join(tools.config.filestore(cr.dbname), path)
		
    def import_moves_from_excel(self, cr, uid, ids, context=None):
        attachment_obj = self.pool.get('ir.attachment')
        attachment_id = attachment_obj.search(cr,uid,[('res_id', '=', ids[0])], context=context)		
        if len(attachment_id)<1: return False

        attach = attachment_obj.browse(cr,uid,attachment_id[0],context=context)
        fname = attach.store_fname
        display_name = attach.name		
        if not fname : return False		
        fname = self.get_full_path(cr, uid, fname)
        wb = load_workbook(filename=fname)	
        #ws = wb.get_sheet_by_name("Sheet1")
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        highest_row = ws.get_highest_row()
        highest_col = ws.get_highest_column()
        title_name = ws.cell(row = 0,column = 0).value	
        title_quantity = ws.cell(row = 0,column = 1).value		

        if highest_col < 2 or title_name != "name" or title_quantity != "quantity":	            		
            raise osv.except_osv(u'Excel错误',u'''文件：%s 格式不正确.''' % display_name)	
			
        row_start = 1
        lines = []	
        product_obj = self.pool.get('product.product')		
        while row_start < highest_row :
            name = ws.cell(row=row_start,column=0).value
            name = name.strip()			
            qty_tmp = ws.cell(row=row_start,column=1)			
            quantity = qty_tmp.get_original_value() or 1
			
            product_ids = product_obj.search(cr, uid, [('name_template','=',name)], context=context)
            if not product_ids : raise osv.except_osv(u'产品名错误',u'''没有产品： %s 。''' % name)			
            lines.append((product_ids[0],quantity))	
			
            row_start += 1			

        picking_obj = self.pool.get('stock.picking').browse(cr,uid,ids[0],context=context)
        picking_type = picking_obj.picking_type_id		
        vals = {	
            'product_id': 0,  
            'product_uom_qty':1, 
            'location_dest_id':picking_type.default_location_dest_id.id, 
            'location_id': picking_type.default_location_src_id.id, 
            'company_id': picking_obj.company_id.id, 
            'date':datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'date_expected':(datetime.datetime.now() + datetime.timedelta(3)).strftime("%Y-%m-%d %H:%M:%S"), 
            'invoice_state':'none', 
            'name':'-', 
            'procure_method':'make_to_stock', 
            'state':'draft',			
            'product_uom':1, 
            'weight_uom_id':1,
            'picking_id': ids[0],			
        }	
        move_obj = self.pool.get('stock.move')		
        for line in lines :
            vals_move = vals.copy()
            vals_move.update({'product_id':line[0], 'product_uom_qty':line[1]})			
            move_obj.create(cr, uid, vals_move, context=context)                		
					
					
    def view_express_data(self, cr, uid, ids, context=None):	
        stock_picking_obj = self.pool.get('stock.picking').browse(cr,uid,ids[0],context=context)	
        if not stock_picking_obj.sale_id : 
            raise osv.except_osv(u'stock.picking 错误',u'''没有销售单与此仓库单有关联''')	
            return False
			
        order_id = 	stock_picking_obj.sale_id.id
        partner_id = stock_picking_obj.partner_id.id 		
        sale_order_line_ids = self.pool.get('sale.order.line').search(cr,uid,[('order_id','=',order_id)],context=context)	
        if len(sale_order_line_ids)< 1: return False		
        eids = self.pool.get('sale.order.line').read(cr,uid,sale_order_line_ids,['coe_no'],context=context)
        express_ids = [ eid['coe_no'] and eid['coe_no'][0] for eid in eids ]			
        if len(express_ids) < 1: 			
            raise osv.except_osv(u'stock.picking 错误',u'''没有快递信息''')		
            return False			
	
        sale_coe_obj = self.pool.get('sale.coe')
        for express_obj in sale_coe_obj.browse(cr,uid,express_ids,context=context):
            if not express_obj.picking_id: 
                express_obj.picking_id = ids[0]		

        mod_obj = self.pool.get('ir.model.data')
        act_obj = self.pool.get('ir.actions.act_window')

        result = mod_obj.get_object_reference(cr, uid, 'loewieec_sync_hk', 'action_loewieec_salecoe')
        id = result and result[1] or False
        result = act_obj.read(cr, uid, [id], context=context)[0]
        result['domain'] = [('id','in', express_ids)]
        result['res_id'] = express_ids
        result['context'] = {'default_sale_id':order_id,'default_customer':partner_id,'default_picking_id':ids[0]}	
		
        return result		
	
    def create_return_lines_from_coe_no(self, cr, uid, ids, context=None):	
		
        picking = self.browse(cr, uid, ids[0], context=context)		
        coenos = picking.note or ''	
        if not coenos : return
        coenos = coenos.strip().split(',')
        coe_list = []		
        for coe in coenos:
            coe = coe.strip()		
            if coe != '' : coe_list.append( coe )
			
        statement = "select s.product_id, s.product_uom_qty, c.name from sale_order_line s left join sale_coe c on s.coe_no=c.id where s.state='done' and s.coe_no in (select id from sale_coe where name in (%s))" %  ("'" + """','""".join(coe_list)	+ "'")		
        cr.execute(statement)
        res = cr.fetchall()		
		
        vals_move = {	
            'create_uid':uid, 		
            'product_id': 0, #, 
            'product_uom_qty':0, 
            'location_dest_id':12, 
            'location_id':9, 
            'company_id':1, 
            'date':datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'date_expected':(datetime.datetime.now() + datetime.timedelta(3)).strftime("%Y-%m-%d %H:%M:%S"), 
            'invoice_state':'none', 
            'name':'-', 
            'procure_method':'make_to_stock', 
            'state':'draft',			
            'product_uom':1, 
            'weight_uom_id':1,
            'picking_id': ids[0],			
        } 
        move_obj = self.pool.get('stock.move')
        for line in res:
            val = vals_move.copy()		
            val.update({'product_id':line[0],'product_uom_qty':line[1],'name':line[2]})	
            move_obj.create(cr,uid,val,context=context)

        return True		
	
    def do_unreserve_no_coe_lines(self, cr, uid, ids, context=None):
		
        picking = self.browse(cr, uid, ids[0], context=context)	
        if picking.state != 'partially_available': return 
		
        quant_obj = self.pool.get("stock.quant")
        move_obj = 	self.pool.get("stock.move")
		
        #waiting_ids = move_obj.search(cr,uid,[('picking_id','=',ids[0]),('state','=','confirmed')],context=context)
        coe_list = []
		
        #for move_unreserved in move_obj.browse(cr,uid,waiting_ids,context=context):
        for move_unreserved in picking.move_lines.filtered(lambda r: r.state == 'confirmed'):		
            if move_unreserved.coe_no not in coe_list :
                coe_list.append(move_unreserved.coe_no)			

        #assigned_ids = move_obj.search(cr,uid,[('picking_id','=',ids[0]),('state','=','assigned')],context=context)					
        #for move in move_obj.browse(cr,uid,assigned_ids,context=context) :
        for move in picking.move_lines.filtered(lambda r: r.state == 'assigned'):	

            if move.coe_no not in coe_list : continue
			
            quant_obj.quants_unreserve(cr, uid, move, context=context)
			
            ancestors = []
            move2 = move
            while move2:
                ancestors += [x.id for x in move2.move_orig_ids]
                move2 = not move2.move_orig_ids and move2.split_from or False
			
            if ancestors:
                move.write({'state': 'waiting'})
            else:
                move.write({'state': 'confirmed'})
	
